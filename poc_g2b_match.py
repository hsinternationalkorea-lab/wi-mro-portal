# -*- coding: utf-8 -*-
"""조달청 100 SKU 매칭 — 정부 조달가 vs 우리 단가 비교"""
import os
import re
import json
import time
import urllib.request
import urllib.parse
from datetime import date, timedelta
from collections import defaultdict
from statistics import mean

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = r"C:\Wholesale Industry\AI Assistant\MRO"
ITEM_MASTER = os.path.join(ROOT, "WI_ItemMaster.xlsx")
OUTPUT_DIR = os.path.join(ROOT, "PriceIntel", "output")

KEY = "566c0b28e67182926927e9beba5ab7ed14d7434902f21341826bb138e386865b"
BASE = "https://apis.data.go.kr/1230000/at/ShoppingMallPrdctInfoService"


def call(path, params):
    p = {"type": "json"}
    p.update(params)
    qs = urllib.parse.urlencode(p, encoding="utf-8")
    url = f"{BASE}{path}?serviceKey={KEY}&{qs}"
    req = urllib.request.Request(url, headers={"User-Agent": "WI-PriceIntel/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        return {"_error": f"HTTP {e.code}"}
    except Exception as e:
        return {"_error": str(e)}


def load_items():
    wb = openpyxl.load_workbook(ITEM_MASTER, data_only=True)
    ws = wb["03_전체마스터"]
    items = []
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code or not str(code).startswith("WH-"):
            continue
        items.append({
            "code": code,
            "category": ws.cell(r, 2).value,
            "name": ws.cell(r, 4).value,
            "spec": ws.cell(r, 5).value,
            "our_price": ws.cell(r, 10).value or 0,
        })
    return items


def make_keyword(item):
    """우리 품목명에서 검색 키워드 추출 (공백 제거된 핵심 명사)"""
    name = str(item["name"] or "")
    # 첫 한글 단어 우선
    korean = re.findall(r"[가-힣]+", name)
    if korean:
        # 가장 긴 한글 단어 선호
        return max(korean, key=len)
    # 한글 없으면 영문 첫 단어
    eng = re.findall(r"[A-Za-z]+", name)
    return eng[0] if eng else name


# 12개월 범위로 광범위 조회
end = date.today() - timedelta(days=1)
start = end - timedelta(days=365)
ds, de = start.strftime("%Y%m%d"), end.strftime("%Y%m%d")


def search_g2b(keyword):
    """다수공급자계약(MAS)에서 분류명 키워드로 검색"""
    return call("/getMASCntrctPrdctInfoList", {
        "numOfRows": 30,
        "pageNo": 1,
        "inqryBgnDt": ds,
        "inqryEndDt": de,
        "prdctClsfcNoNm": keyword,  # 분류명 (broader)
    })


def parse_results(api_response):
    if "_error" in api_response:
        return [], api_response["_error"]
    body = api_response.get("response", {}).get("body", {})
    items_raw = body.get("items", [])
    if not isinstance(items_raw, list):
        return [], None
    parsed = []
    for it in items_raw:
        try:
            price = int(it.get("cntrctPrceAmt", 0) or 0)
        except (ValueError, TypeError):
            price = 0
        parsed.append({
            "g2b_price": price,
            "spec": it.get("prdctSpecNm", ""),
            "maker": it.get("prdctMakrNm", ""),
            "corp": it.get("cntrctCorpNm", ""),
            "unit": it.get("prdctUnit", ""),
            "image": it.get("prdctImgUrl", ""),
            "lead_time": it.get("dlvrTmlmtDaynum", ""),
            "classification": it.get("prdctClsfcNoNm", ""),
            "origin": it.get("prdctOrgplceNm", ""),
        })
    return parsed, None


def main():
    print("=" * 70)
    print(f"조달청 100 SKU 매칭 (기간: {ds}~{de})")
    print("=" * 70)

    items = load_items()
    results = []

    for i, item in enumerate(items, 1):
        kw = make_keyword(item)
        if not kw or len(kw) < 2:
            results.append({**item, "keyword": kw, "matches": [], "error": "키워드 짧음"})
            continue

        api = search_g2b(kw)
        matches, err = parse_results(api)

        results.append({**item, "keyword": kw, "matches": matches, "error": err})
        if i % 10 == 0:
            print(f"   ... {i}/{len(items)}")
        time.sleep(0.3)

    # 통계
    matched = [r for r in results if r["matches"] and not r.get("error")]
    print(f"\n매칭률: {len(matched)}/{len(results)} = {len(matched)/len(results)*100:.1f}%")

    # 가격 비교 (매칭 시 G2B 최저가 vs 우리 단가)
    diffs = []
    for r in matched:
        prices = [m["g2b_price"] for m in r["matches"] if m["g2b_price"] > 0]
        if not prices or not r["our_price"]:
            continue
        g2b_min = min(prices)
        diff = (r["our_price"] - g2b_min) / r["our_price"] * 100
        diffs.append(diff)

    if diffs:
        print(f"\n가격 차이 (양수=우리 우위, 음수=조달이 더 쌈):")
        print(f"  평균:   {mean(diffs):+.1f}%")
        cheap = [d for d in diffs if d > 0]
        expen = [d for d in diffs if d < 0]
        print(f"  우리가 싼 SKU:  {len(cheap)}개")
        print(f"  우리가 비싼 SKU: {len(expen)}개")

    # 카테고리별 매칭률
    by_cat = defaultdict(lambda: {"total": 0, "matched": 0})
    for r in results:
        c = r["category"]
        by_cat[c]["total"] += 1
        if r["matches"]:
            by_cat[c]["matched"] += 1
    print("\n카테고리별 매칭:")
    for cat, s in sorted(by_cat.items()):
        rate = s["matched"] / s["total"] * 100
        print(f"  {cat}: {s['matched']}/{s['total']} ({rate:.0f}%)")

    # 엑셀 출력
    out = os.path.join(OUTPUT_DIR, f"poc_g2b_{date.today().strftime('%Y%m%d')}.xlsx")
    write_excel(results, out)
    print(f"\n결과: {out}")


def write_excel(results, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "조달청_매칭"

    HF = PatternFill("solid", fgColor="0089D0")
    HFONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)

    headers = [
        "WI 품번", "카테고리", "품목명", "규격/PN", "검색어",
        "우리 단가", "조달 최저", "차이 % (양수=우리우위)", "매칭 수",
        "최저가 사양", "제조사", "공급사", "단위", "원산지", "분류",
        "납품일수", "이미지URL",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HF
        cell.font = HFONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, item in enumerate(results, 2):
        ws.cell(r, 1, item["code"])
        ws.cell(r, 2, item["category"])
        ws.cell(r, 3, item["name"])
        ws.cell(r, 4, item["spec"])
        ws.cell(r, 5, item.get("keyword", ""))
        ws.cell(r, 6, item["our_price"])
        if item["matches"]:
            prices = [m["g2b_price"] for m in item["matches"] if m["g2b_price"] > 0]
            if prices and item["our_price"]:
                g2b_min = min(prices)
                best = min(item["matches"], key=lambda x: x["g2b_price"] if x["g2b_price"] else 9e9)
                diff = (item["our_price"] - g2b_min) / item["our_price"] * 100
                ws.cell(r, 7, g2b_min)
                ws.cell(r, 8, round(diff, 1))
                ws.cell(r, 9, len(item["matches"]))
                ws.cell(r, 10, best["spec"])
                ws.cell(r, 11, best["maker"])
                ws.cell(r, 12, best["corp"])
                ws.cell(r, 13, best["unit"])
                ws.cell(r, 14, best["origin"])
                ws.cell(r, 15, best["classification"])
                ws.cell(r, 16, best["lead_time"])
                ws.cell(r, 17, best["image"])
                # 색상
                if diff < -10:
                    ws.cell(r, 8).fill = PatternFill("solid", fgColor="FBEEEC")
                elif diff > 10:
                    ws.cell(r, 8).fill = PatternFill("solid", fgColor="E8F4EC")

    widths = [12, 8, 24, 22, 12, 12, 12, 14, 8, 35, 16, 16, 6, 10, 18, 8, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results)+1}"
    wb.save(path)


if __name__ == "__main__":
    main()
