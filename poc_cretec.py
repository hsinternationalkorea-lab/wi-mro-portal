# -*- coding: utf-8 -*-
"""
크레텍 크롤링 PoC — Playwright 기반
- robots.txt: User-agent: * Allow: / Crawl-delay: 20
- 100 SKU 매칭 시도
- 결과: 우리 단가 vs 크레텍 가격 비교
"""
import os
import re
import time
from datetime import date
from collections import defaultdict
from statistics import mean

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

ROOT = r"C:\Wholesale Industry\AI Assistant\MRO"
ITEM_MASTER = os.path.join(ROOT, "WI_ItemMaster.xlsx")
OUTPUT_DIR = os.path.join(ROOT, "PriceIntel", "output")

CRAWL_DELAY = 20  # robots.txt 명시
SAMPLE_SIZE = 10  # PoC: 우선 10개만 (10×20초 = 3분 30초)


def load_items():
    wb = openpyxl.load_workbook(ITEM_MASTER, data_only=True)
    ws = wb["03_전체마스터"]
    items = []
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code or not str(code).startswith("WH-"):
            continue
        items.append({
            "code": code,
            "category": ws.cell(r, 2).value,
            "name": ws.cell(r, 4).value,
            "spec": ws.cell(r, 5).value,
            "our_price": ws.cell(r, 10).value or 0,
        })
    return items


def make_keyword(item):
    """검색 키워드: 첫 한글 단어 (가장 식별성 높음)"""
    name = str(item["name"] or "")
    korean = re.findall(r"[가-힣]+", name)
    if korean:
        return max(korean, key=len)
    eng = re.findall(r"[A-Za-z]+", name)
    return eng[0] if eng else name


def search_cretec(page, keyword):
    """크레텍 검색 후 결과 추출"""
    # 검색 URL 패턴 시도 (실제 URL은 사이트 진입 후 검증)
    page.goto("https://www.cretec.kr/", timeout=20000, wait_until="domcontentloaded")
    page.wait_for_timeout(2000)

    # 검색 input 찾기 (다양한 선택자 시도)
    selectors = [
        "input[type='search']",
        "input[name='keyword']",
        "input[placeholder*='검색']",
        "input.search",
        "#search_input",
        "header input",
    ]
    found = False
    for sel in selectors:
        try:
            if page.locator(sel).first.is_visible(timeout=1500):
                page.locator(sel).first.fill(keyword)
                page.locator(sel).first.press("Enter")
                found = True
                break
        except Exception:
            continue

    if not found:
        return {"error": "검색 input 못 찾음", "html_sample": page.content()[:500]}

    # 결과 페이지 로딩 대기
    page.wait_for_timeout(3000)

    # 가격 텍스트 추출
    body = page.content()
    prices = re.findall(r"(\d{1,3}(?:,\d{3})+)\s*원", body)
    prices_int = sorted(set(int(p.replace(",", "")) for p in prices if int(p.replace(",", "")) > 100))

    # 상품명 후보
    titles = []
    for sel in ["a.product-name", ".product-title", ".item-name", "h3 a", ".prd_name"]:
        try:
            elements = page.locator(sel).all()
            for el in elements[:5]:
                titles.append(el.inner_text())
            if titles:
                break
        except Exception:
            continue

    return {
        "url": page.url,
        "prices": prices_int[:10],
        "titles": titles[:5],
        "raw_count": len(prices),
    }


def main():
    items = load_items()[:SAMPLE_SIZE]  # 샘플
    print(f"=== 크레텍 크롤링 PoC ({len(items)}개 SKU, Crawl-delay {CRAWL_DELAY}초) ===\n")

    results = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            locale="ko-KR",
            viewport={"width": 1280, "height": 800},
        )

        for i, item in enumerate(items, 1):
            kw = make_keyword(item)
            print(f"[{i}/{len(items)}] {item['code']} '{kw}' 검색 중...")

            page = context.new_page()
            try:
                r = search_cretec(page, kw)
                if "error" in r:
                    print(f"   ⚠ {r['error']}")
                    results.append({**item, "kw": kw, **r})
                else:
                    print(f"   ✅ URL: {r['url'][:80]}")
                    print(f"      가격 {len(r['prices'])}개, 타이틀 {len(r['titles'])}개")
                    if r["prices"]:
                        print(f"      가격 샘플: {r['prices'][:3]}")
                    if r["titles"]:
                        print(f"      타이틀 샘플: {r['titles'][0][:50]}")
                    results.append({**item, "kw": kw, **r})
            except Exception as e:
                print(f"   ❌ {type(e).__name__}: {str(e)[:80]}")
                results.append({**item, "kw": kw, "error": str(e)})
            finally:
                page.close()

            if i < len(items):
                print(f"   ⏳ Crawl-delay {CRAWL_DELAY}초 대기...")
                time.sleep(CRAWL_DELAY)

        browser.close()

    # 결과 보고
    matched = [r for r in results if r.get("prices")]
    print(f"\n매칭률: {len(matched)}/{len(results)} ({len(matched)/len(results)*100:.0f}%)")
    print(f"\n결과:")
    for r in results:
        if r.get("prices"):
            min_p = min(r["prices"])
            print(f"  {r['code']} '{r['kw']}': {len(r['prices'])}개 가격, 최저 {min_p:,}원 (우리 {r['our_price']:,}원)")


if __name__ == "__main__":
    main()
