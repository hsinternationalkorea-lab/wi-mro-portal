# -*- coding: utf-8 -*-
"""
Mouser Search API PoC
- 100 SKU 중 전자부품·전기 카테고리에서 Mouser 매칭 시도
- Mouser는 전자부품 전문 (저항·콘덴서·IC·LED·커넥터·케이블)
- 우리 ItemMaster의 E 카테고리(전기조명) + 전기 관련 부품 시도
"""
import os
import json
import time
import urllib.request
from datetime import date

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = r"C:\Wholesale Industry\AI Assistant\MRO"
ITEM_MASTER = os.path.join(ROOT, "WI_ItemMaster.xlsx")
OUTPUT_DIR = os.path.join(ROOT, "PriceIntel", "output")
ENV_FILE = os.path.join(ROOT, "PriceIntel", "_secrets", ".env")

MOUSER_ENDPOINT = "https://api.mouser.com/api/v1/search/keyword"


def load_env():
    env = {}
    with open(ENV_FILE, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


def mouser_search(keyword, api_key, records=10):
    """Mouser Search API — keyword search"""
    url = f"{MOUSER_ENDPOINT}?apiKey={api_key}"
    body = {
        "SearchByKeywordRequest": {
            "keyword": keyword,
            "records": records,
            "startingRecord": 0,
            "searchOptions": "InStock",  # 재고있는 것만
            "searchWithYourSignUpLanguage": "false",
        }
    }
    data = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json", "Accept": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception as e:
        return {"error": str(e)}


def parse_mouser_part(part):
    """Mouser 응답에서 핵심 정보 추출"""
    price_breaks = part.get("PriceBreaks", [])
    # PriceBreaks: [{"Quantity":1, "Price":"$0.123", "Currency":"USD"}, ...]
    min_price_usd = None
    if price_breaks:
        try:
            # 수량 1개 가격 (가장 비싼 단가)
            qty1 = next((pb for pb in price_breaks if pb.get("Quantity") == 1), price_breaks[0])
            price_str = qty1.get("Price", "")
            min_price_usd = float(price_str.replace("$", "").replace(",", ""))
        except Exception:
            pass

    return {
        "mouser_pn": part.get("MouserPartNumber", ""),
        "mfr": part.get("Manufacturer", ""),
        "mfr_pn": part.get("ManufacturerPartNumber", ""),
        "description": part.get("Description", ""),
        "category": part.get("Category", ""),
        "availability": part.get("Availability", ""),
        "image": part.get("ImagePath", ""),
        "price_usd": min_price_usd,
        "url": part.get("ProductDetailUrl", ""),
    }


# ────────────────────── 첫 번째 테스트: API 동작 확인 ──────────────────────

def test_api():
    env = load_env()
    api_key = env.get("MOUSER_API_KEY", "")
    if not api_key:
        print("[ERROR] MOUSER_API_KEY 없음")
        return

    print("=" * 60)
    print("Mouser API 연결 테스트")
    print("=" * 60)
    print(f"API Key: {api_key[:8]}...{api_key[-4:]}")
    print()

    # 표준 부품으로 테스트 (저항)
    test_kw = "10K resistor 0805"
    print(f"테스트 검색: '{test_kw}'")
    result = mouser_search(test_kw, api_key, records=5)

    if result.get("error"):
        print(f"❌ 오류: {result['error']}")
        return False

    if "Errors" in result and result["Errors"]:
        print(f"❌ Mouser 오류: {result['Errors']}")
        return False

    search_results = result.get("SearchResults", {})
    parts = search_results.get("Parts", [])
    total = search_results.get("NumberOfResult", 0)
    print(f"✅ 응답 수신. 총 {total}개 결과 중 {len(parts)}개 반환")
    print()

    if parts:
        print("샘플 결과:")
        for p in parts[:3]:
            parsed = parse_mouser_part(p)
            print(f"  {parsed['mouser_pn']:<15} | {parsed['mfr']:<20} | "
                  f"${parsed['price_usd']:.4f}/ea | {parsed['description'][:50]}")
    return True


# ────────────────────── 100 SKU 매칭 시도 ──────────────────────

def load_items():
    wb = openpyxl.load_workbook(ITEM_MASTER, data_only=True)
    ws = wb["03_전체마스터"]
    items = []
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code or not str(code).startswith("WH-"):
            continue
        items.append({
            "code": code,
            "category": ws.cell(r, 2).value,
            "name": ws.cell(r, 4).value,
            "spec": ws.cell(r, 5).value,
            "our_price": ws.cell(r, 10).value or 0,
        })
    return items


def make_mouser_query(item):
    """Mouser 검색어 생성 — 영문/모델번호 우선"""
    name = str(item["name"] or "")
    spec = str(item["spec"] or "")
    full = f"{name} {spec}"

    # 영문·숫자만 추출 (Mouser는 한글 검색 의미 없음)
    import re
    eng_only = " ".join(re.findall(r"[A-Za-z0-9\-/]+", full))
    return eng_only.strip()[:50]


def main_full_run():
    env = load_env()
    api_key = env.get("MOUSER_API_KEY", "")

    print("=" * 70)
    print(f"Mouser API 100 SKU 매칭 시도 — {date.today()}")
    print("=" * 70)

    items = load_items()
    results = []

    print(f"\n[검색 진행] (Mouser API rate: 30/min, 0.5초 간격)")
    for i, item in enumerate(items, 1):
        query = make_mouser_query(item)

        if not query or len(query) < 3:
            results.append({
                **item, "query": query, "matched": False,
                "skip_reason": "영문 키워드 없음 (한글만)",
                "parts": [],
            })
            continue

        result = mouser_search(query, api_key, records=5)

        if result.get("error") or "Errors" in result and result.get("Errors"):
            results.append({
                **item, "query": query, "matched": False,
                "skip_reason": f"API 오류: {result.get('error') or result.get('Errors')}",
                "parts": [],
            })
            continue

        parts_raw = result.get("SearchResults", {}).get("Parts", [])
        parts = [parse_mouser_part(p) for p in parts_raw]

        results.append({
            **item, "query": query, "matched": len(parts) > 0,
            "skip_reason": "",
            "parts": parts,
        })

        if i % 10 == 0:
            print(f"   ... {i}/{len(items)}")
        time.sleep(0.5)  # rate limit (30/min)

    # 통계
    matched = [r for r in results if r["matched"]]
    print()
    print(f"매칭 성공: {len(matched)}/{len(results)} ({len(matched)/len(results)*100:.1f}%)")
    print()

    # 카테고리별 매칭률
    by_cat = {}
    for r in results:
        cat = r["category"]
        by_cat.setdefault(cat, {"total": 0, "matched": 0})
        by_cat[cat]["total"] += 1
        if r["matched"]:
            by_cat[cat]["matched"] += 1
    print("카테고리별 매칭:")
    for cat, s in sorted(by_cat.items()):
        rate = s["matched"] / s["total"] * 100
        print(f"  {cat}: {s['matched']}/{s['total']} ({rate:.0f}%)")

    # 엑셀 출력
    out = os.path.join(OUTPUT_DIR, f"poc_mouser_{date.today().strftime('%Y%m%d')}.xlsx")
    write_excel(results, out)
    print(f"\n결과: {out}")


def write_excel(results, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mouser_매칭"

    HF = PatternFill("solid", fgColor="0089D0")
    HFONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)

    headers = [
        "WI 품번", "카테고리", "품목명", "규격/PN", "검색어 (영문)",
        "우리 단가 (KRW)", "매칭 수", "Mouser PN", "제조사", "MPN",
        "Description", "단가 (USD)", "재고", "URL", "비고",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HF
        cell.font = HFONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, item in enumerate(results, 2):
        ws.cell(r, 1, item["code"])
        ws.cell(r, 2, item["category"])
        ws.cell(r, 3, item["name"])
        ws.cell(r, 4, item["spec"])
        ws.cell(r, 5, item.get("query", ""))
        ws.cell(r, 6, item["our_price"])
        ws.cell(r, 7, len(item["parts"]) if item.get("parts") else 0)
        if item["parts"]:
            top = item["parts"][0]
            ws.cell(r, 8, top["mouser_pn"])
            ws.cell(r, 9, top["mfr"])
            ws.cell(r, 10, top["mfr_pn"])
            ws.cell(r, 11, top["description"][:60])
            ws.cell(r, 12, top["price_usd"])
            ws.cell(r, 13, top["availability"])
            ws.cell(r, 14, top["url"])
        ws.cell(r, 15, item.get("skip_reason", ""))

    widths = [12, 8, 24, 20, 22, 12, 8, 14, 16, 16, 30, 10, 10, 30, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "test":
        test_api()
    else:
        # 먼저 연결 테스트
        if test_api():
            print("\n" + "=" * 70)
            print("연결 OK. 100 SKU 검색 진행...")
            print("=" * 70 + "\n")
            time.sleep(2)
            main_full_run()
