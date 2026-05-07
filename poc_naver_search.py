# -*- coding: utf-8 -*-
"""
Phase 0 PoC: ItemMaster 100 SKU vs 네이버쇼핑 가격 비교
- 입력: WI_ItemMaster.xlsx (캐노니컬, 100 SKU)
- API: 네이버 검색 API (쇼핑) — 일 25,000회 무료
- 출력: PriceIntel/output/poc_price_comparison_YYYYMMDD.xlsx + 통계 보고서

핵심 검증 질문:
  1. 100 SKU 중 외부 검색에서 매칭 가능한 비율은?
  2. 우리 단가 vs 외부 최저가 차이의 분포는?
  3. 카테고리별로 차이가 다른가?
  4. 사업성: 검색엔진 가치가 통계적으로 유의한가?
"""
import os
import re
import time
import json
import urllib.request
import urllib.parse
from datetime import datetime, date
from collections import defaultdict
from statistics import mean, median, stdev

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ────────────────────── 설정 ──────────────────────
ROOT = r"C:\Wholesale Industry\AI Assistant\MRO"
ITEM_MASTER = os.path.join(ROOT, "WI_ItemMaster.xlsx")
OUTPUT_DIR = os.path.join(ROOT, "PriceIntel", "output")
ENV_FILE = os.path.join(ROOT, "PriceIntel", "_secrets", ".env")

# 환경변수 로드 (.env 파일 단순 파서)
def load_env():
    if not os.path.exists(ENV_FILE):
        print(f"[ERROR] .env 파일 없음: {ENV_FILE}")
        print(f"        .env.example 을 .env 로 복사 + API 키 입력")
        return None
    env = {}
    with open(ENV_FILE, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env

# ────────────────────── 1. ItemMaster 로드 ──────────────────────
def load_items():
    wb = openpyxl.load_workbook(ITEM_MASTER, data_only=True)
    ws = wb["03_전체마스터"]
    items = []
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code or not str(code).startswith("WH-"):
            continue
        items.append({
            "code": code,
            "category": ws.cell(r, 2).value,
            "category_name": ws.cell(r, 3).value,
            "name": ws.cell(r, 4).value,
            "spec": ws.cell(r, 5).value,
            "our_price": ws.cell(r, 10).value or 0,
        })
    return items

# ────────────────────── 2. 네이버 검색 ──────────────────────
def naver_search(query, client_id, client_secret, display=10):
    """네이버 쇼핑 검색 API"""
    enc = urllib.parse.quote(query)
    url = f"https://openapi.naver.com/v1/search/shop.json?query={enc}&display={display}&sort=asc"
    req = urllib.request.Request(url)
    req.add_header("X-Naver-Client-Id", client_id)
    req.add_header("X-Naver-Client-Secret", client_secret)
    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            if response.status == 200:
                return json.loads(response.read().decode("utf-8"))
            return None
    except Exception as e:
        return {"error": str(e)}

# ────────────────────── 3. 검색어 생성 ──────────────────────
def make_query(item):
    """ItemMaster 행 → 네이버 검색어 (정제)"""
    name = str(item["name"] or "").strip()
    spec = str(item["spec"] or "").strip()
    # HTML 태그/특수문자 제거
    name = re.sub(r"<[^>]+>", "", name)
    spec = re.sub(r"<[^>]+>", "", spec)
    # 너무 긴 규격은 잘라냄
    if len(spec) > 25:
        spec = spec[:25]
    if spec and spec.lower() != "none":
        return f"{name} {spec}"
    return name

# ────────────────────── 4. 가격 추출 ──────────────────────
def extract_prices(naver_result):
    """네이버 응답에서 가격 리스트 추출"""
    if not naver_result or "items" not in naver_result:
        return []
    prices = []
    for it in naver_result["items"]:
        try:
            lp = int(it.get("lprice", 0))  # 최저가
            if lp > 0:
                prices.append({
                    "price": lp,
                    "title": re.sub(r"<[^>]+>", "", it.get("title", "")),
                    "mall": it.get("mallName", ""),
                    "link": it.get("link", ""),
                    "image": it.get("image", ""),
                })
        except Exception:
            continue
    return prices

# ────────────────────── 5. 메인 ──────────────────────
def main():
    env = load_env()
    if not env:
        return
    cid = env.get("NAVER_CLIENT_ID", "")
    csec = env.get("NAVER_CLIENT_SECRET", "")
    if "여기에" in cid or not cid or not csec:
        print(f"[ERROR] .env 파일에 실제 API 키 입력 필요")
        return

    print(f"[1/4] ItemMaster 로드")
    items = load_items()
    print(f"      → {len(items)} SKU")

    print(f"[2/4] 네이버 검색 API 호출 (각 SKU 0.2초 간격)")
    results = []
    for i, item in enumerate(items, 1):
        query = make_query(item)
        api_result = naver_search(query, cid, csec, display=10)
        prices = extract_prices(api_result)

        if prices:
            min_p = min(p["price"] for p in prices)
            max_p = max(p["price"] for p in prices)
            avg_p = int(mean(p["price"] for p in prices))
            top_match = prices[0]
        else:
            min_p = max_p = avg_p = 0
            top_match = {"title": "", "mall": "", "link": "", "image": ""}

        our = item["our_price"]
        diff_pct = ((our - min_p) / our * 100) if our and min_p else 0

        results.append({
            "code": item["code"],
            "category": item["category"],
            "name": item["name"],
            "spec": item["spec"],
            "query": query,
            "our_price": our,
            "naver_min": min_p,
            "naver_avg": avg_p,
            "naver_max": max_p,
            "diff_pct": diff_pct,
            "match_count": len(prices),
            "top_title": top_match["title"][:60],
            "top_mall": top_match["mall"],
            "top_link": top_match["link"],
        })

        if i % 10 == 0:
            print(f"      ... {i}/{len(items)}")
        time.sleep(0.2)  # API 부담 방지

    print(f"[3/4] 통계 분석 + 엑셀 출력")

    # 매칭률
    matched = [r for r in results if r["match_count"] > 0]
    match_rate = len(matched) / len(results) * 100

    # 가격 차이 (음수=우리가 비싸다, 양수=우리가 싸다)
    diffs = [r["diff_pct"] for r in matched if r["our_price"] > 0]
    cheaper = [d for d in diffs if d > 0]   # 우리가 더 싸다
    expensive = [d for d in diffs if d < 0]  # 우리가 더 비싸다

    # 카테고리별
    by_cat = defaultdict(list)
    for r in matched:
        if r["our_price"] > 0:
            by_cat[r["category"]].append(r["diff_pct"])

    # 엑셀 출력
    out_file = os.path.join(
        OUTPUT_DIR,
        f"poc_price_comparison_{date.today().strftime('%Y%m%d')}.xlsx"
    )
    write_excel(results, out_file)

    # 콘솔 보고서
    print()
    print("=" * 70)
    print(f"Phase 0 PoC 결과 — {date.today().strftime('%Y-%m-%d')}")
    print("=" * 70)
    print(f"총 SKU: {len(results)}개")
    print(f"매칭 성공: {len(matched)}개 ({match_rate:.1f}%)")
    print(f"매칭 실패: {len(results) - len(matched)}개")
    print()
    if diffs:
        print(f"가격 차이 분포 (양수 = 우리가 싸다):")
        print(f"  평균: {mean(diffs):+.1f}%")
        print(f"  중앙값: {median(diffs):+.1f}%")
        if len(diffs) > 1:
            print(f"  표준편차: {stdev(diffs):.1f}%")
        print(f"  최대(우리 우위): {max(diffs):+.1f}%")
        print(f"  최소(우리 열위): {min(diffs):+.1f}%")
        print()
        print(f"우리가 더 싼 SKU: {len(cheaper)}개 ({len(cheaper)/len(diffs)*100:.0f}%)")
        print(f"우리가 더 비싼 SKU: {len(expensive)}개 ({len(expensive)/len(diffs)*100:.0f}%)")
        print()
        print(f"카테고리별 평균 차이:")
        for cat, ds in sorted(by_cat.items()):
            print(f"  {cat}: {mean(ds):+.1f}% (n={len(ds)})")
    print()
    print(f"[4/4] 결과 엑셀: {out_file}")


def write_excel(results, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "비교결과"

    HEADER_FILL = PatternFill("solid", fgColor="0089D0")
    HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    NUM_FONT = Font(name="맑은 고딕", size=10)

    headers = [
        "WI 품번", "카테고리", "품목명", "규격/PN", "검색어",
        "우리 단가", "네이버 최저", "네이버 평균", "네이버 최고",
        "차이 % (양수=우리우위)", "매칭 수", "최저가 상품명", "판매몰", "링크",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, item in enumerate(results, 2):
        ws.cell(r, 1, item["code"])
        ws.cell(r, 2, item["category"])
        ws.cell(r, 3, item["name"])
        ws.cell(r, 4, item["spec"])
        ws.cell(r, 5, item["query"])
        ws.cell(r, 6, item["our_price"])
        ws.cell(r, 7, item["naver_min"])
        ws.cell(r, 8, item["naver_avg"])
        ws.cell(r, 9, item["naver_max"])
        ws.cell(r, 10, round(item["diff_pct"], 1) if item["match_count"] > 0 else None)
        ws.cell(r, 11, item["match_count"])
        ws.cell(r, 12, item["top_title"])
        ws.cell(r, 13, item["top_mall"])
        ws.cell(r, 14, item["top_link"])

        # 색상 (우리가 비싸면 빨강, 싸면 초록)
        diff = item["diff_pct"]
        if item["match_count"] > 0:
            if diff < -10:
                ws.cell(r, 10).fill = PatternFill("solid", fgColor="FBEEEC")  # 빨간계
            elif diff > 10:
                ws.cell(r, 10).fill = PatternFill("solid", fgColor="E8F4EC")  # 초록계

    # 컬럼 너비
    widths = [12, 8, 28, 22, 32, 12, 12, 12, 12, 16, 8, 35, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results)+1}"

    wb.save(path)


if __name__ == "__main__":
    main()
