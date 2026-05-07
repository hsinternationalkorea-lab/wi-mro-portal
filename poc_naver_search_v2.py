# -*- coding: utf-8 -*-
"""
Phase 0.5 PoC v2 — 매칭 정확도 개선
- 검수 데이터 기반 개선: 53% → 목표 80%+
- 핵심 변경:
  1. 검색어 정제: 모델번호·브랜드 우선 추출
  2. 다중 검색 전략: 모델번호 단독 → 모델+카테고리 → 풀텍스트 fallback
  3. 결과 필터링: 제목 유사도 (Jaccard) 임계값 적용
  4. 단위 정규화: "1매당 가격" 자동 환산 시도
"""
import os
import re
import time
import json
import urllib.request
import urllib.parse
from datetime import date
from collections import defaultdict
from statistics import mean, median

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = r"C:\Wholesale Industry\AI Assistant\MRO"
ITEM_MASTER = os.path.join(ROOT, "WI_ItemMaster.xlsx")
OUTPUT_DIR = os.path.join(ROOT, "PriceIntel", "output")
ENV_FILE = os.path.join(ROOT, "PriceIntel", "_secrets", ".env")


def load_env():
    env = {}
    with open(ENV_FILE, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def load_items():
    wb = openpyxl.load_workbook(ITEM_MASTER, data_only=True)
    ws = wb["03_전체마스터"]
    items = []
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code or not str(code).startswith("WH-"):
            continue
        items.append({
            "code": code,
            "category": ws.cell(r, 2).value,
            "name": ws.cell(r, 4).value,
            "spec": ws.cell(r, 5).value,
            "our_price": ws.cell(r, 10).value or 0,
        })
    return items


# ────────────────────── 핵심 v2: 모델번호/브랜드 추출 ──────────────────────

# 모델번호 패턴 (영문대문자+숫자+하이픈 조합 등)
MODEL_PATTERNS = [
    r"\b[A-Z]{2,}[-_]?[\d]+[A-Z\d\-]*\b",        # TR-632E, BPK-HD, PB212LH
    r"\b[A-Z]+\d{3,}[A-Z\-\d]*\b",                # E39, M5
    r"\b\d+[Mm][Mm]\b",                           # 300mm
    r"\b[A-Z][A-Z]\-\d+\b",                       # WW-2309
    r"\b\d+\.?\d*\s*(?:mm|cm|mil|inch|m|L)\b",   # 8mil, 1L
]

# 알려진 브랜드 (확장 가능)
BRANDS = [
    "3M", "PHILIPS", "필립스", "DAILOVE", "Honeywell", "BW",
    "스마토", "TYGON", "케이엠", "FOCUS", "Versum",
]


def extract_models(text):
    """텍스트에서 모델번호·브랜드 추출"""
    text = str(text or "")
    found = set()
    for pat in MODEL_PATTERNS:
        for m in re.findall(pat, text):
            if len(m) >= 3:
                found.add(m.strip())
    for brand in BRANDS:
        if brand.lower() in text.lower():
            found.add(brand)
    return list(found)


def make_query_v2(item):
    """검수 결과 기반 v2 검색어 — 모델번호 우선, 일반명사만일 땐 카테고리 추가"""
    name = str(item["name"] or "").strip()
    spec = str(item["spec"] or "").strip()
    full_text = f"{name} {spec}"

    models = extract_models(full_text)

    # 전략 1: 모델번호가 있으면 모델 + 핵심 명사 1개
    if models:
        # 가장 긴 모델 (가장 식별성 높음)
        primary_model = max(models, key=len)
        # 짧은 핵심 명사 (한글) — 너무 일반적인 건 제외
        generic_words = {"호스", "장갑", "캡", "비닐", "케이블", "타이", "전구"}
        words = re.findall(r"[가-힣]+", name)
        # 4자 이상 한글 단어 또는 비-일반 단어 우선
        keywords = [w for w in words if len(w) >= 3 and w not in generic_words]
        if keywords:
            return f"{primary_model} {keywords[0]}"
        return primary_model

    # 전략 2: 모델번호 없으면 풀 키워드 + 카테고리 보강
    cat_hint = {
        "C": "크린룸",
        "S": "안전",
        "L": "실험실",
        "T": "공구",
        "M": "의료",
    }.get(item["category"], "")

    if spec and spec.lower() != "none":
        q = f"{name} {spec}"
    else:
        q = name
    # 너무 짧으면 카테고리 보강
    if len(q) < 8 and cat_hint:
        q = f"{cat_hint} {q}"
    return q[:60]


# ────────────────────── 결과 필터링: 제목 유사도 ──────────────────────

def normalize_text(s):
    s = str(s or "").lower()
    s = re.sub(r"<[^>]+>", "", s)
    s = re.sub(r"[^\w가-힣]+", " ", s)
    return s


def jaccard(a, b):
    """Jaccard 유사도 (단어 수준)"""
    sa = set(normalize_text(a).split())
    sb = set(normalize_text(b).split())
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def filter_by_similarity(items_naver, our_text, threshold=0.15):
    """검색결과 중 우리 품목과 단어 겹침 일정 수준 이상만 통과"""
    out = []
    for it in items_naver:
        sim = jaccard(our_text, it.get("title", ""))
        if sim >= threshold:
            it["_sim"] = sim
            out.append(it)
    return out


# ────────────────────── 단위 정규화 ──────────────────────

UNIT_HINTS = [
    (r"(\d+)\s*매", "매"),
    (r"(\d+)\s*개", "개"),
    (r"(\d+)\s*pcs?", "pcs"),
    (r"(\d+)\s*Box", "box"),
    (r"(\d+)\s*set", "set"),
    (r"(\d+)\s*ea", "ea"),
]


def extract_unit_count(text):
    """제목에서 '50매', '100개' 같은 수량 단위 추출. 없으면 1"""
    text = str(text or "").lower()
    for pat, _ in UNIT_HINTS:
        m = re.search(pat, text)
        if m:
            try:
                n = int(m.group(1))
                if n > 1:
                    return n
            except Exception:
                pass
    return 1


# ────────────────────── 네이버 API ──────────────────────

def naver_search(query, cid, csec, display=10):
    enc = urllib.parse.quote(query)
    url = f"https://openapi.naver.com/v1/search/shop.json?query={enc}&display={display}&sort=asc"
    req = urllib.request.Request(url)
    req.add_header("X-Naver-Client-Id", cid)
    req.add_header("X-Naver-Client-Secret", csec)
    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            if response.status == 200:
                return json.loads(response.read().decode("utf-8"))
    except Exception:
        pass
    return None


def extract_prices(naver_result):
    if not naver_result or "items" not in naver_result:
        return []
    out = []
    for it in naver_result["items"]:
        try:
            lp = int(it.get("lprice", 0))
            if lp > 0:
                out.append({
                    "price": lp,
                    "title": re.sub(r"<[^>]+>", "", it.get("title", "")),
                    "mall": it.get("mallName", ""),
                    "link": it.get("link", ""),
                })
        except Exception:
            continue
    return out


# ────────────────────── 메인 ──────────────────────

def main():
    env = load_env()
    cid = env.get("NAVER_CLIENT_ID")
    csec = env.get("NAVER_CLIENT_SECRET")

    print("[1/4] ItemMaster 로드 (100 SKU)")
    items = load_items()

    print("[2/4] v2 정제 검색 (모델번호 우선 + 유사도 필터)")
    results = []
    for i, item in enumerate(items, 1):
        query = make_query_v2(item)
        api = naver_search(query, cid, csec, display=15)
        prices = extract_prices(api)

        # 우리 품목과 유사도 필터
        our_text = f'{item["name"]} {item["spec"]}'
        filtered = filter_by_similarity(prices, our_text, threshold=0.15)

        # 단위 정규화: 우리 단가가 박스라 가정하고 1개당 환산
        # (이건 추정이므로 별도 컬럼으로 보고)
        if filtered:
            # 단위당 가격 환산
            for p in filtered:
                cnt = extract_unit_count(p["title"])
                p["per_unit_price"] = p["price"] / cnt
                p["unit_count"] = cnt

            min_p = min(p["price"] for p in filtered)
            min_per_unit = min(p["per_unit_price"] for p in filtered)
            avg_p = int(mean(p["price"] for p in filtered))
            best = min(filtered, key=lambda x: x["price"])
        else:
            min_p = avg_p = 0
            min_per_unit = 0
            best = {"title": "", "mall": "", "link": "", "_sim": 0}

        our = item["our_price"]
        diff_pct = ((our - min_p) / our * 100) if our and min_p else 0

        results.append({
            "code": item["code"],
            "category": item["category"],
            "name": item["name"],
            "spec": item["spec"],
            "query_v2": query,
            "our_price": our,
            "naver_min": min_p,
            "naver_avg": avg_p,
            "naver_per_unit": int(min_per_unit) if min_per_unit else 0,
            "diff_pct": diff_pct,
            "match_count": len(filtered),
            "raw_match_count": len(prices),
            "top_title": best.get("title", "")[:60],
            "top_mall": best.get("mall", ""),
            "top_link": best.get("link", ""),
            "top_sim": round(best.get("_sim", 0), 2),
        })

        if i % 10 == 0:
            print(f"      ... {i}/{len(items)}")
        time.sleep(0.2)

    print("[3/4] 통계 + 엑셀 출력")

    matched = [r for r in results if r["match_count"] > 0]
    diffs = [r["diff_pct"] for r in matched if r["our_price"] > 0]

    out_file = os.path.join(
        OUTPUT_DIR, f"poc_v2_{date.today().strftime('%Y%m%d')}.xlsx"
    )
    write_excel(results, out_file)

    print()
    print("=" * 70)
    print(f"Phase 0.5 v2 결과 — {date.today().strftime('%Y-%m-%d')}")
    print("=" * 70)
    print(f"총 SKU: {len(results)}개")
    print(f"v1 매칭률: 89% → v2 매칭률 (유사도 필터 적용): {len(matched)/len(results)*100:.1f}%")
    print()
    if diffs:
        print(f"가격 차이 (필터 후, 양수=우리 우위):")
        print(f"  평균:   {mean(diffs):+.1f}%")
        print(f"  중앙값: {median(diffs):+.1f}%")
        cheap = [d for d in diffs if d > 0]
        expensive = [d for d in diffs if d < 0]
        print(f"  우리가 싼 SKU:  {len(cheap)}개 ({len(cheap)/len(diffs)*100:.0f}%)")
        print(f"  우리가 비싼 SKU: {len(expensive)}개 ({len(expensive)/len(diffs)*100:.0f}%)")
    print()
    print(f"[4/4] 결과: {out_file}")


def write_excel(results, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "v2_비교결과"

    HF = PatternFill("solid", fgColor="0089D0")
    HFONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)

    headers = [
        "WI 품번", "카테고리", "품목명", "규격/PN", "검색어 v2",
        "우리 단가", "네이버 최저", "네이버 평균", "1개당 환산",
        "차이 % (양수=우리우위)", "매칭 수 (필터 전)", "유사도",
        "최저가 상품명", "판매몰", "링크",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HF
        cell.font = HFONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, item in enumerate(results, 2):
        ws.cell(r, 1, item["code"])
        ws.cell(r, 2, item["category"])
        ws.cell(r, 3, item["name"])
        ws.cell(r, 4, item["spec"])
        ws.cell(r, 5, item["query_v2"])
        ws.cell(r, 6, item["our_price"])
        ws.cell(r, 7, item["naver_min"])
        ws.cell(r, 8, item["naver_avg"])
        ws.cell(r, 9, item["naver_per_unit"])
        ws.cell(r, 10, round(item["diff_pct"], 1) if item["match_count"] > 0 else None)
        ws.cell(r, 11, f'{item["match_count"]} / {item["raw_match_count"]}')
        ws.cell(r, 12, item["top_sim"])
        ws.cell(r, 13, item["top_title"])
        ws.cell(r, 14, item["top_mall"])
        ws.cell(r, 15, item["top_link"])

        diff = item["diff_pct"]
        if item["match_count"] > 0:
            if diff < -10:
                ws.cell(r, 10).fill = PatternFill("solid", fgColor="FBEEEC")
            elif diff > 10:
                ws.cell(r, 10).fill = PatternFill("solid", fgColor="E8F4EC")

    widths = [12, 8, 28, 22, 26, 12, 12, 12, 12, 16, 12, 8, 35, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results)+1}"
    wb.save(path)


if __name__ == "__main__":
    main()
