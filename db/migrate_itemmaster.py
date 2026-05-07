# -*- coding: utf-8 -*-
"""
ItemMaster (100 SKU) → Supabase products 테이블 마이그레이션
- 우리 직거래 카탈로그를 DB에 시드
- is_directly_sold = TRUE
- is_published = TRUE (즉시 포털 노출)
- 가격: cost_price 부재시 source_price 사용, list_price = cost × margin
"""
import os
import openpyxl

ROOT = r"C:\Wholesale Industry\AI Assistant\MRO"
ITEM_MASTER = os.path.join(ROOT, "WI_ItemMaster.xlsx")

# 카테고리별 마진 (schema.sql과 동일)
CATEGORY_MARGIN = {
    "M": 25.0, "F": 18.0, "L": 22.0, "C": 20.0,
    "S": 25.0, "E": 15.0, "P": 18.0, "T": 22.0, "O": 20.0,
}


def load_master():
    wb = openpyxl.load_workbook(ITEM_MASTER, data_only=True)
    ws = wb["03_전체마스터"]
    rows = []
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code or not str(code).startswith("WH-"):
            continue
        rows.append({
            "wi_code": code,
            "category_code": ws.cell(r, 2).value,
            "category_name": ws.cell(r, 3).value,
            "name_ko": ws.cell(r, 4).value,
            "spec": ws.cell(r, 5).value,
            "first_requester": ws.cell(r, 6).value,
            "supply_count": ws.cell(r, 7).value or 0,
            "cumulative_qty": ws.cell(r, 8).value or 0,
            "cumulative_amt_thousand": ws.cell(r, 9).value or 0,
            "last_unit_price": ws.cell(r, 10).value or 0,
            "first_supply_date": ws.cell(r, 11).value,
            "status": ws.cell(r, 12).value,
        })
    return rows


def to_sql_inserts(items):
    """SQL INSERT 문 생성 (Supabase SQL Editor에 붙여넣기 가능)"""
    lines = [
        "-- ItemMaster 100 SKU → products 마이그레이션",
        "-- Supabase SQL Editor에서 실행",
        "",
    ]

    for item in items:
        cat = item["category_code"]
        cost = item["last_unit_price"] or 0
        margin = CATEGORY_MARGIN.get(cat, 20.0)
        list_price = round(cost * (1 + margin / 100))

        # SQL Escape
        def esc(v):
            if v is None:
                return "NULL"
            if isinstance(v, (int, float)):
                return str(v)
            s = str(v).replace("'", "''")
            return f"'{s}'"

        keywords = f"{item['name_ko']} {item['spec'] or ''} {item['category_name'] or ''}"
        keywords = keywords.replace("'", "")

        sql = f"""INSERT INTO products (
    wi_code, source_code, source_product_id,
    category_code, name_ko, spec,
    search_keywords,
    cost_price, list_price, margin_pct,
    is_directly_sold, is_published, quality_score
) VALUES (
    {esc(item['wi_code'])}, 'wi_master', {esc(item['wi_code'])},
    {esc(cat)}, {esc(item['name_ko'])}, {esc(item['spec'])},
    {esc(keywords[:500])},
    {cost}, {list_price}, {margin},
    TRUE, TRUE, 1.00
);"""
        lines.append(sql)

    lines.append("")
    lines.append(f"-- 총 {len(items)} 건 삽입")
    return "\n".join(lines)


def main():
    items = load_master()
    print(f"로드: {len(items)} SKU")

    sql = to_sql_inserts(items)
    out = os.path.join(os.path.dirname(__file__), "seed_itemmaster.sql")
    with open(out, "w", encoding="utf-8") as f:
        f.write(sql)
    print(f"SQL 파일 생성: {out}")
    print(f"라인 수: {sql.count(chr(10))}")

    # 통계
    by_cat = {}
    for it in items:
        c = it["category_code"]
        by_cat.setdefault(c, 0)
        by_cat[c] += 1
    print("\n카테고리별 분포:")
    for c, cnt in sorted(by_cat.items()):
        m = CATEGORY_MARGIN.get(c, 20.0)
        print(f"  {c}: {cnt}건 (마진 {m}%)")


if __name__ == "__main__":
    main()
